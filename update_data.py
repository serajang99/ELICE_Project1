import openpyxl
from flask_apscheduler import APScheduler
from models import Book
from db_connect import db
import os
import natsort
from datetime import datetime
from flask import Flask

scheduler = APScheduler()


app = Flask(__name__)

app.config.from_mapping(
    SECRET_KEY='dev',
)
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:1234@localhost:3306/elice_library'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)


@scheduler.task('interval', id='update_data', seconds=10)
def update_data():
    with app.app_context():
        wb = openpyxl.load_workbook(filename='data/book.xlsx')
        ws = wb.active

        img_dir = os.listdir('./static/img/book_img')
        img_list = natsort.natsorted(img_dir)

        for r in range(2, ws.max_row):

            if ws.cell(r, 2).value is None:
                break

            title = ws.cell(r, 2).value
            publisher = ws.cell(r, 3).value
            author = ws.cell(r, 4).value
            publication_date = ws.cell(r, 5).value
            pages = ws.cell(r, 6).value
            isbn = ws.cell(r, 7).value
            description = ws.cell(r, 8).value
            link = ws.cell(r, 9).value
            img_value = img_list[r-1]
            img = f'img/book_img/{img_value}'
            data = {
                'title': title,
                'publisher': publisher,
                'author': author,
                'publication_date': publication_date,
                'pages': pages,
                'isbn': isbn,
                'description': description,
                'link': link,
                'img': img,
                'register_time': datetime.fromisoformat('2000-01-01 00:00:00'),
            }
            book = Book(data)

            db.session.add(book)
            db.session.commit()

        print('데이터 저장 완료')
        db.session.close()


def init_scheduler(app):
    scheduler.api_enabled = True
    scheduler.init_app(app)
    scheduler.start()


update_data()
